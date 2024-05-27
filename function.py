from fastapi import FastAPI
from fastapi.responses import FileResponse
# from pydantic import BaseModel
from database import DB_Connect
from sqlalchemy import text
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side


app = FastAPI()


@app.get("/get_emp_details")
def get_emp_details(name: str = None):
    conn = None
    try:
        with DB_Connect() as conn:
            if name:
                query = text("SELECT student_code,student_name,sem_no,SUM(semester_marks) AS sem_total FROM master_student_marks AS msm INNER JOIN master_student AS ms ON ms.id = msm.student_id INNER JOIN master_semester AS sem ON sem.id = msm.sem_id INNER JOIN master_subject AS sub ON sub.id = msm.subject_id GROUP BY msm.student_id,msm.sem_id,msm.subject_idWHERE student_name = :name")
                result = conn.execute(query, {"student_name": name}).mappings().all()
            else:
                query = "SELECT student_code,student_name,sem_no,SUM(semester_marks) AS sem_total FROM master_student_marks AS msm INNER JOIN master_student AS ms ON ms.id = msm.student_id INNER JOIN master_semester AS sem ON sem.id = msm.sem_id INNER JOIN master_subject AS sub ON sub.id = msm.subject_id GROUP BY msm.student_id,msm.sem_id,msm.subject_id;"
                result = conn.execute(text(query)).mappings().all()
        return result
    except Exception as e:
        print(f"Error executing query: {e}")
        return None
    finally:
        if conn:
            conn.close()


@app.get("/export_to_excel")
def export_to_excel():
    try:
        data = get_emp_details()
        wb = Workbook()
        ws = wb.active

        headers = list(data[0].keys())
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_index, value=header)

        for row_index, row_data in enumerate(data, start=2):
            for col_index,cell_value in enumerate(row_data.values(), start=1):
                ws.cell(row=row_index, column=col_index, value=cell_value)

        file_path = 'sheet1.xlsx'
        wb.save(file_path)
        return FileResponse(file_path, filename='sheet1.xlsx')
    except Exception as e:
        return {"error": f"{e}"}
    
    
template_path = "default_template.xlsx"   
@app.get('/default_template')   
def Default_template():
    try:
        wb = Workbook()
        ws = wb.active
        headers = ["student_code","student_name","sem_no","sem_total"]
        bold_font = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="3366FF", end_color="3366FF", fill_type="solid")
        border = Border(left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))
        
        for col_index, header in enumerate(headers, start=1):
            cell=ws.cell(row=1, column=col_index, value=header)
            cell.font = bold_font
            cell.fill = fill
            cell.border = border

        file_path = "default_template.xlsx" 
        wb.save(file_path)

    except Exception as e:
        return {"error": f"{e}"}
    
def Insert_data_to_default_template():
    try:
        data = get_emp_details()
        print("Retrieved data:", data) 
        wb = load_workbook(template_path)
        ws = wb.active
        fill = PatternFill(start_color="FF99CC", end_color="FF99CC", fill_type="solid")
        border = Border(left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))
        for row_index, row_data in enumerate(data, start=2):
            for col_index,cell_value in enumerate(row_data.values(), start=1):
                cell=ws.cell(row=row_index, column=col_index, value=cell_value)
                cell.fill = fill
                cell.border = border

        wb.save(template_path)

        return FileResponse(template_path, filename='default_template.xlsx')
    except Exception as e:
        return {"error": f"{e}"}
    
@app.post("/insert_data/")
def insert_data():
    try:
        Insert_data_to_default_template()
        print("data inserted")
        return {"message": "Data inserted into the template successfully"}
    except Exception as e:
        return {"error": f"{e}"}
    

# class Record(BaseModel):
#     name : str
#     department : str
#     username :str
#     password :str
#     subject : str
#     date : date

# @app.post("/post_emp")
# def Post(name: Annotated[str, Form()], department: Annotated[str, Form()],username: Annotated[str, Form()], password: Annotated[str, Form()],department: Annotated[str, Form()], subject: Annotated[str, Form()],date: Annotated[date, Form()], name: Annotated[str, Form()],department: Annotated[str, Form()],):
#     try:
#         with DB_Connect() as conn:
#             query = "INSERT INTO employee (name, department) VALUES (:name, :department); INSERT INTO admin (name, username , password) VALUES (:username, :password); INSERT INTO staff (name, department ,subject  , date) VALUES (:name, :department ,:subject ,:date); INSERT INTO HOD (name, department) VALUES (:name, :department);"
#             conn.execute(text(query), {"name": name, "department": department})
#             conn.commit()
#             return {"message": "Record created successfully"}
#     except Exception as e:
#         print(f"Error: {e}")
#     finally:
#         if conn:
#             conn.close()



